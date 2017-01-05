#! python3
#monthlyFinanceBoard.py - automates the monthly Finance & Manager board proccess



#Importing modules
import openpyxl
import sys
import os
import re
from openpyxl.utils import column_index_from_string


#Couldn't think of the name - where the Created, Modified, Approved By columns are stored
columnLocation=['Created Column Location','Modified Column Location','Approved Column Location']
#The columns in numbers
inNumbers=[1,2,3]
#List for what rows the empty cells are located
emptyRows=[]
#Regex expression for date
dateObject = re.compile(r'\d\d\d\d-\d\d-\d\d')
#Placeholder for the formname
formName ='placeholder'


def changeDate(): 
#Change the date formatting 
	#Iterates through the row/columns and change the value of the cell to first 10 characters if it matches the pattern dateObject
	for rowNum in range(1,sheet.max_row+1):
		for colNum in range(1,sheet.max_column):
			unchangedDate = str(sheet.cell(row=rowNum,column=colNum).value)
			if dateObject.search(unchangedDate):
				sheet.cell(row=rowNum,column=colNum).value = unchangedDate[:10]				
def calculateManagerDifference(): 
#Calculates difference for Manager
	#Calculated Manager value will be at max column + 1 
	calculatedManager=sheet.max_column+1
	type = 'Manager'

	#Making Manager Sheet by copying data. Skips rows that has empty value for 'Approved'
	managerSheet= wb.create_sheet('Manager')
	counter = 0 #Counts how many rows to skip idk how it works it just does
	for rowNum in range(1,sheet.max_row+1):
		if rowNum not in emptyRows:
			for colNum in range(1, sheet.max_column):
				managerSheet.cell(row=rowNum-counter,column=colNum).value = sheet.cell(row=rowNum,column=colNum).value
		else:
			for colNum in range(1, sheet.max_column):
				managerSheet.cell(row=rowNum-counter,column=colNum).value = sheet.cell(row=rowNum+1,column=colNum).value
			counter = counter+1
	
	#Adding Calculated Manager Column
	managerSheet.cell(row=1,column=sheet.max_column+1).value= "Calculated Manager"
	
	#Writing in the formula for manager calculation
	for rowNum in range(2,sheet.max_row - counter):
		managerSheet.cell(row=rowNum,column=calculatedManager).value= '='+columnLocation[2]+':'+columnLocation[2]+'-'+columnLocation[0]+':'+columnLocation[0]
	
	findAverage(managerSheet,type)
def calculateFinanceDifference(): 
#Calculates difference for Finance
	#Making Finance Sheet. Copies over data. If Approved cell is empty, then replace it with value from 'Created' column
	calculatedFinance=sheet.max_column+1
	type = 'Finance'
	financeSheet= wb.create_sheet('Finance')
	for rowNum in range(1,sheet.max_row+1):
		if rowNum in emptyRows:
			for colNum in range(1, sheet.max_column):
				financeSheet.cell(row=rowNum,column=colNum).value = sheet.cell(row=rowNum,column=colNum).value
				financeSheet.cell(row=rowNum,column=inNumbers[2]).value = sheet.cell(row=rowNum,column=inNumbers[0]).value
		else:
			for colNum in range(1, sheet.max_column):
				financeSheet.cell(row=rowNum,column=colNum).value = sheet.cell(row=rowNum,column=colNum).value
	
	#Adding Calculated Manager Column
	financeSheet.cell(row=1,column=sheet.max_column+1).value= "Calculated Finance"
	
	#Doing the actual calculation
	for rowNum in range(2,sheet.max_row):
		financeSheet.cell(row=rowNum,column=calculatedFinance).value= '='+columnLocation[1]+':'+columnLocation[1]+'-'+columnLocation[2]+':'+columnLocation[2]
	
	findAverage(financeSheet,type)
def findColumnsAndEmptyRows(): 
#Find the last column & also find the rows that have empty values - for different calculations
	#Iterate through columns and find matching values.
	for colNum in range(1,sheet.max_column+1):
		search = str(sheet.cell(row=1,column=colNum).value)
		if search == "Created":
			columnLocation[0] = str(sheet.cell(row=1,column=colNum).column)
			inNumbers[0] = column_index_from_string(columnLocation[0])
		elif search == "Modified":
			columnLocation[1] = str(sheet.cell(row=1,column=colNum).column)
			inNumbers[1] = column_index_from_string(columnLocation[1])
		elif search == "Approved By Date":
			columnLocation[2] = str(sheet.cell(row=1,column=colNum).column)
			inNumbers[2] = column_index_from_string(columnLocation[2])
		elif search == "Approval Date":
			columnLocation[2] = str(sheet.cell(row=1,column=colNum).column)
			inNumbers[2] = column_index_from_string(columnLocation[2])
		elif search == "Approved Date":
			columnLocation[2] = str(sheet.cell(row=1,column=colNum).column)
			inNumbers[2] = column_index_from_string(columnLocation[2])
			
	
	#Find rows where Approved Date is blank
	for rowNum in range(1, sheet.max_row+1):
		if sheet.cell(row=rowNum,column=inNumbers[2]).value == None:
			emptyRows.append(sheet.cell(row=rowNum,column=inNumbers[2]).row)
def findAverage(sheetname,type): 
#Find the average and output it as a cell
	sheetname.cell(row=2,column=sheetname.max_column+1).value='=AVERAGE('+sheet.cell(row=2,column=sheetname.max_column).column+':'+sheet.cell(row=2,column=sheetname.max_column).column+')'
	sheetname.cell(row=1,column=sheetname.max_column).value= 'Average'
def whichForm(formName): 
#Asks the name of the form
	#Too lazy to make a switch statement so I'm just going to use if-else statements.
	keepGoing = True
	while keepGoing == True:
		choice = input("""Please indicate the form that you are using.
					1: Form 1
					2: Form 2
					3: Form 3
					4: Form 4""")
		if int(choice) == 1:
			formName = "Form 1"
			keepGoing=False
			return formName
		elif int(choice) == 2:
			formName= "Form 2"
			keepGoing=False
			return formName
		elif int(choice) == 3:
			formName = "Form 3"
			keepGoing=False
			return formName
		elif int(choice) == 4:
			formName = "Form 4"
			keepGoing=False
			return formName
		else:
			print('\nPlease input a valid value. Input 1,2,3, or 4.\n')
def outputNumbers(): 
#Writes the number of forms to a file
	outputFile = open('Number of forms.txt','a')
	outputFile.write('\n'+form +' has ' +str(sheet.max_row)+' items')
	outputFile.close

#Omitted - openpyxl does not do the actual calculations and excel document must be opened to get the calculated value.
#This defeats the purpose of 'automation' so I just left this function out
"""
def outputIntoFile(type):
	wb2 = openpyxl.load_workbook(saveTo,data_only=True)
	print(wb2.get_sheet_names())
	sheetname = wb2[type]
	print(wb2.active)
	outputFile = open('Output.txt','a')
	if type == 'Finance':
		outputFile.write('\n'+form +' Finance Difference: '+ str(sheetname.cell(row=2,column=sheetname.max_column).value))
		outputFile.close()
	elif type == 'Manager':
		outputFile.write('\n'+form +' Manager Difference: '+str(sheetname.cell(row=2,column=sheetname.max_column).internal_value))
		outputFile.close()
"""


print("\nPlease put the script in the same folder as the excel spreadsheets - it will not work otherwise")					
loadFrom = input('Please input the name of the excel workbook(only the title, CAse SenSATive):')+'.xlsx'
location = os.path.join(sys.path[0], loadFrom)

wb = openpyxl.load_workbook(location) #Opening the worksheet
sheet = wb.active #Opening active sheet

form = whichForm(formName)
changeDate()
findColumnsAndEmptyRows()
saveTo = input('Please input the name of the output spreadsheet file:')+'.xlsx'
calculateManagerDifference()
calculateFinanceDifference()
outputNumbers()
wb.save(saveTo)

print('Done')
finish =input("Press \'Enter\' to exit")

